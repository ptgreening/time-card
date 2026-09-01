"""Compute each formula's result and store it as the cell's cached value.

LibreOffice cannot recalculate in this environment, and openpyxl writes formulas
with no cached value — so every formula cell would read as empty to anything that
does not recalculate. Excel recalculates on open regardless, but caching the
values means the file is correct the moment it is opened anywhere, and it lets the
results be verified here rather than taken on trust.
"""
import zipfile, shutil, re, datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

SRC = 'Haunted-Trail-Hours-Aug-2026.xlsx'
wb = load_workbook(SRC)

EPOCH = dt.datetime(1899, 12, 30)
def serial(v):
    if isinstance(v, dt.datetime): return (v - EPOCH).total_seconds()/86400
    if isinstance(v, dt.date):     return (dt.datetime(v.year,v.month,v.day) - EPOCH).days
    return v

def num(ws, ref):
    v = ws[ref].value
    if v is None or v == '': return None
    if isinstance(v, str) and v.startswith('='): return CALC.get((ws.title, ref))
    return serial(v)

CALC = {}

def compliance(net, meals, deduct, mins, began):
    if net is None or net <= 5: return ''
    if meals == 0 and deduct == 0:
        return ('WAIVER REQUIRED - no meal, 5-6 hr shift' if net <= 6
                else 'VIOLATION - no meal taken, over 5 hrs')
    if meals == 0 and deduct > 0: return 'REVIEW - lunch button, no times recorded'
    if mins is not None and mins < 30: return 'VIOLATION - meal under 30 minutes'
    if began is not None and began > 5: return 'VIOLATION - meal began after 5th hour'
    if net > 12 and meals < 2: return 'VIOLATION - no 2nd meal, over 12 hrs'
    if net > 10 and meals < 2: return 'WAIVER REQUIRED - no 2nd meal, 10-12 hrs'
    return ''

for ws in wb:
    if ws.title in ('Read Me', 'Needs Review'): continue
    for row in range(1, ws.max_row + 1):
        if ws.title == 'All Punches':
            if isinstance(ws.cell(row,6).value, str) and ws.cell(row,6).value.startswith('='):
                d, o = num(ws,f'D{row}'), num(ws,f'E{row}')
                CALC[(ws.title, f'F{row}')] = (o - d) * 24
            continue
        # pay period sheets
        h = ws.cell(row,8).value
        if isinstance(h, str) and h.startswith('=IF(E'):
            D_,E_,F_,G_ = (num(ws,f'{c}{row}') for c in 'DEFG')
            gross = (G_-D_)*24 if E_ is None else (E_-D_)*24 + (G_-F_)*24
            CALC[(ws.title, f'H{row}')] = gross
            ded = ws.cell(row,9).value or 0
            net = gross - ded
            CALC[(ws.title, f'J{row}')] = net
            meals = ws.cell(row,11).value or 0
            mins  = None if E_ is None else (F_-E_)*1440
            began = None if E_ is None else (E_-D_)*24
            CALC[(ws.title, f'L{row}')] = '' if mins  is None else mins
            CALC[(ws.title, f'M{row}')] = '' if began is None else began
            txt = compliance(net, meals, ded, mins, began)
            CALC[(ws.title, f'N{row}')] = txt
            CALC[(ws.title, f'O{row}')] = 1 if txt.startswith('VIOLATION') else 0
            CALC[(ws.title, f'P{row}')] = 'OT' if net > 8 else ''
        # totals rows
        for col in (8,9,10,15):
            v = ws.cell(row,col).value
            m = isinstance(v,str) and re.fullmatch(r'=SUM\((\w)(\d+):\w(\d+)\)', v)
            if m:
                L,a,b = m.group(1), int(m.group(2)), int(m.group(3))
                CALC[(ws.title, f'{get_column_letter(col)}{row}')] = sum(
                    x for r in range(a,b+1)
                    if isinstance(x := CALC.get((ws.title,f'{L}{r}')), (int,float)))
        v = ws.cell(row,16).value
        if isinstance(v,str) and v.startswith('=IF(J') and 'OVER 40' in v:
            CALC[(ws.title, f'P{row}')] = 'OVER 40' if (CALC.get((ws.title,f'J{row}')) or 0) > 40 else ''

# ── write the values into the sheet XML ──
NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
order = {ws.title: i+1 for i, ws in enumerate(wb)}
shutil.copy(SRC, 'tmp.xlsx')
zin = zipfile.ZipFile('tmp.xlsx')
zout = zipfile.ZipFile(SRC, 'w', zipfile.ZIP_DEFLATED)
name_by_idx = {v: k for k, v in order.items()}
injected = 0
for item in zin.infolist():
    data = zin.read(item.filename)
    m = re.fullmatch(r'xl/worksheets/sheet(\d+)\.xml', item.filename)
    if m:
        title = name_by_idx.get(int(m.group(1)))
        xml = data.decode('utf-8')
        def fix(cm):
            global injected
            ref, attrs, f_tag = cm.group(1), cm.group(2), cm.group(3)
            val = CALC.get((title, ref))
            if val is None or val == '':
                return cm.group(0)          # no result, or an empty-string result
            injected += 1
            attrs = re.sub(r'\s+t="[^"]*"', '', attrs)   # drop any existing type
            if isinstance(val, str):
                esc = val.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
                return f'<c r="{ref}"{attrs} t="str">{f_tag}<v>{esc}</v></c>'
            return f'<c r="{ref}"{attrs}>{f_tag}<v>{val!r}</v></c>'
        xml = re.sub(r'<c r="([A-Z]+\d+)"([^>]*)>(<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?</c>',
                     fix, xml, flags=re.S)
        data = xml.encode('utf-8')
    zout.writestr(item, data)
zout.close(); zin.close()
print(f'cached {injected} formula results')
