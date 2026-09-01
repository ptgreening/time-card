import json, datetime as dt
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = json.load(open('days.json'))
P = json.load(open('punches.json'))
iso = lambda s: dt.datetime.fromisoformat(s) if s else None
for r in D:
    r['date']=dt.date.fromisoformat(r['date'])
    for k in ('clock_in','clock_out','meal_out','meal_in'): r[k]=iso(r[k])

ARIAL   = 'Arial'
HDR_F   = PatternFill('solid', fgColor='1F3864')
TOT_F   = PatternFill('solid', fgColor='D9E2F3')
FLAG_F  = PatternFill('solid', fgColor='FCE4E4')
HOLD_F  = PatternFill('solid', fgColor='FFF2CC')
TITLE   = Font(name=ARIAL, size=14, bold=True)
HDR     = Font(name=ARIAL, size=10, bold=True, color='FFFFFF')
BASE    = Font(name=ARIAL, size=10)
BOLD    = Font(name=ARIAL, size=10, bold=True)
INPUT   = Font(name=ARIAL, size=10, color='0000FF')
MUTED   = Font(name=ARIAL, size=9, color='808080')
RED     = Font(name=ARIAL, size=10, bold=True, color='C00000')
thin    = Side(style='thin', color='BFBFBF')
BOX     = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = Workbook(); wb.remove(wb.active)

def head(ws, cols, row, widths):
    for c,(t,w) in enumerate(zip(cols,widths),1):
        cell=ws.cell(row,c,t); cell.font=HDR; cell.fill=HDR_F
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=BOX
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[row].height=30

# ── READ ME ───────────────────────────────────────────────────
ws = wb.create_sheet('Read Me')
ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=30
ws.column_dimensions['C'].width=86
ws['B2']='Haunted Trail — Payroll Hours Report'; ws['B2'].font=TITLE
ws['B3']=f'Generated {dt.date.today():%B %d, %Y} · Pay periods run Tuesday through Monday'; ws['B3'].font=MUTED
r=5
def note(label, text, f=BASE):
    global r
    ws.cell(r,2,label).font=BOLD
    c=ws.cell(r,3,text); c.font=f; c.alignment=Alignment(wrap_text=True,vertical='top')
    ws.row_dimensions[r].height=max(14, 13*(len(text)//95+1)); r+=1

note('READ THIS FIRST','80% of all recorded hours in this period come from shifts where someone forgot to clock '
     'out — including spans of 4 and 8 days. Those are quarantined on the "Needs Review" tab and are NOT included '
     'in any total. Correct them before paying anyone.', RED)
r+=1
note('Hours format','Decimal, to two places. 7.50 means seven hours thirty minutes.')
note('Net hours','Gross time on the clock, minus any unpaid meal break.')
note('Meal — clocked out','Room Escape, Axe Throw and Build Crew clock out for lunch. That time sits between '
     '"Meal Out" and "Meal In" and is already excluded from Gross, so Meal Deduct is 0.00.')
note('Meal — lunch button','Manager, Actor, Makeup Artist, Crowd Control and Carnival Crew press a button instead '
     'of clocking out. That records a 0.50 deduction, shown in the Meal Deduct column.')
note('Day assignment','A shift belongs to the day it STARTED. A shift running past midnight stays on the '
     'day it began.')
r+=1
ws.cell(r,2,'FLAGS').font=Font(name=ARIAL,size=11,bold=True); r+=1
note('Meal break rules','California Labor Code 512 and the IWC Wage Orders. A meal period must be at '
     'least 30 uninterrupted minutes, and must BEGIN before the end of the 5th hour of work.')
note('  Over 5 hours','A meal period is required. If none was taken, that is a violation.')
note('  5 to 6 hours','The meal may be waived by mutual consent. Shown as WAIVER REQUIRED — not a penalty, but you '
     'need a signed waiver on file to rely on it.')
note('  Over 10 hours','A second meal period is required. Between 10 and 12 hours it may be waived by mutual '
     'consent, provided the first meal was actually taken.')
note('  Over 12 hours','The second meal can no longer be waived. Missing it is a violation.')
note('Penalty Hrs','One hour of pay at the regular rate per workday, under Labor Code 226.7. Capped at one hour '
     'per day for meal violations no matter how many occur that day.')
note('REVIEW rows','The lunch button records THAT a meal was taken but not WHEN or for HOW LONG, so compliance '
     'cannot be verified from the data. See the note at the bottom of this page.')
note('OT DAY','More than 8.00 net hours in a single day.')
note('OVER 40','More than 40.00 net hours in a pay period, per employee.')
note('HELD','Excluded from totals pending correction. See the Needs Review tab.')
r+=1
note('Blue figures','Entered from the time clock records, not calculated. Everything black is a formula.')
note('Punch times','Clock In, Meal Out, Meal In and Clock Out are shown on every row, so any flagged row can be '
     'checked against the punches behind it.')
note('A row showing 6.00 can still flag MBP','Hours are displayed to two places but the flags test the exact time '
     'worked. Kevin Tran on 08/15 reads 6.00 because he worked 6 hours and 9 seconds. The flag is correct; the '
     'display is simply rounded. Erring toward flagging is deliberate — an extra look costs less than a missed penalty.',
     Font(name=ARIAL, size=10, italic=True))
r+=1
ws.cell(r,2,'GAP TO FIX').font=Font(name=ARIAL,size=11,bold=True,color='C00000'); r+=1
note('The lunch button records no time','Roles that press the lunch button instead of clocking out record only '
     'that a meal happened — no start time and no duration. California compliance turns on WHEN the meal began '
     'and HOW LONG it lasted, so those days cannot be shown compliant from the records. They are marked REVIEW '
     'rather than assumed good. Recording a timestamp when the button is pressed would close this.', RED)
r+=1
note('Rest breaks are not tracked','A 10-minute paid rest period is required per 4 hours worked, and carries its '
     'own separate one-hour penalty. The time clock does not record rest breaks, so no rest violations appear '
     'here. That does not mean there were none.', RED)
ws.sheet_view.showGridLines=False

# ── PAY PERIOD SHEETS ─────────────────────────────────────────
def pstart(d): return d - dt.timedelta(days=(d.weekday()-1)%7)
periods=defaultdict(list)
for x in D: periods[pstart(x['date'])].append(x)

COLS=['Employee','Date','Day','Clock In','Meal Out','Meal In','Clock Out',
      'Gross Hrs','Meal Deduct','Net Hrs','Meals','Meal Mins','Meal Began\n(hrs into shift)',
      'Meal Break Compliance (CA)','Penalty Hrs','OT DAY','Notes']
W=[22,11,6,11,11,11,11,9,9,9,7,9,11,48,9,8,30]
NCOL=17

for ps in sorted(periods):
    pe=ps+dt.timedelta(days=6)
    ws=wb.create_sheet(f'{ps:%b %d} - {pe:%b %d}')
    ws['A1']=f'Pay Period: Tuesday {ps:%B %d} through Monday {pe:%B %d, %Y}'; ws['A1'].font=TITLE
    head(ws,COLS,3,W)
    ws.freeze_panes='A4'
    row=4
    byemp=OrderedDict()
    for x in sorted(periods[ps], key=lambda z:(z['name'],z['date'])):
        byemp.setdefault(x['name'],[]).append(x)

    for name, recs in byemp.items():
        first=row
        netcells=[]
        for x in recs:
            held=x['review']
            ws.cell(row,1,name).font=BASE
            c=ws.cell(row,2,x['date']); c.font=BASE; c.number_format='mm/dd/yyyy'
            ws.cell(row,3,f"{x['date']:%a}").font=BASE
            for col,key in ((4,'clock_in'),(5,'meal_out'),(6,'meal_in'),(7,'clock_out')):
                v=x[key]
                if v:
                    c=ws.cell(row,col,v); c.font=INPUT; c.number_format='h:mm AM/PM'
                    # A punch that landed on a later date needs to say so.
                    if v.date()!=x['date']:
                        c.number_format='mm/dd h:mm AM/PM'
            if held:
                for col in range(1,NCOL+1): ws.cell(row,col).fill=HOLD_F
                ws.cell(row,8,'HELD').font=RED
                ws.cell(row,14,'Not assessed - shift needs correction').font=MUTED
                ws.cell(row,NCOL,x['why']).font=BASE
            else:
                g=f'=IF(E{row}="",(G{row}-D{row})*24,(E{row}-D{row})*24+(G{row}-F{row})*24)'
                ws.cell(row,8,g).font=BASE
                ws.cell(row,9,0.5 if x['button_lunch'] else 0).font=INPUT
                ws.cell(row,10,f'=H{row}-I{row}').font=BOLD
                ws.cell(row,11,1 if x['meal_out'] else 0).font=INPUT
                ws.cell(row,12,f'=IF(E{row}="","",(F{row}-E{row})*1440)').font=BASE
                ws.cell(row,13,f'=IF(E{row}="","",(E{row}-D{row})*24)').font=BASE
                # Labor Code 512: meal required over 5 hrs, must BEGIN before the end of
                # the 5th hour, at least 30 uninterrupted minutes. 5-6 hrs may be waived;
                # a 2nd meal is required over 10 and unwaivable over 12.
                ws.cell(row,14,
                    f'=IF(J{row}<=5,"",'
                    f'IF(AND(K{row}=0,I{row}=0),IF(J{row}<=6,"WAIVER REQUIRED - no meal, 5-6 hr shift",'
                    f'"VIOLATION - no meal taken, over 5 hrs"),'
                    f'IF(AND(K{row}=0,I{row}>0),"REVIEW - lunch button, no times recorded",'
                    f'IF(L{row}<30,"VIOLATION - meal under 30 minutes",'
                    f'IF(M{row}>5,"VIOLATION - meal began after 5th hour",'
                    f'IF(AND(J{row}>12,K{row}<2),"VIOLATION - no 2nd meal, over 12 hrs",'
                    f'IF(AND(J{row}>10,K{row}<2),"WAIVER REQUIRED - no 2nd meal, 10-12 hrs","")))))))').font=BASE
                ws.cell(row,15,f'=IF(LEFT(N{row},9)="VIOLATION",1,0)').font=RED
                ws.cell(row,16,f'=IF(J{row}>8,"OT","")').font=RED
                if x['note']: ws.cell(row,NCOL,x['note']).font=MUTED
                netcells.append(f'J{row}')
            for col in (8,9,10,12,13,15):
                ws.cell(row,col).number_format='0.00'
                ws.cell(row,col).alignment=Alignment(horizontal='center')
            for col in (3,11,16): ws.cell(row,col).alignment=Alignment(horizontal='center')
            ws.cell(row,14).alignment=Alignment(horizontal='left')
            for col in range(1,NCOL+1): ws.cell(row,col).border=BOX
            row+=1
        # employee total for the pay period
        rng=f'J{first}:J{row-1}'
        ws.cell(row,1,f'TOTAL — {name}').font=BOLD
        ws.cell(row,10,f'=SUM({rng})').font=BOLD
        ws.cell(row,10).number_format='0.00'
        ws.cell(row,10).alignment=Alignment(horizontal='center')
        ws.cell(row,8,f'=SUM(H{first}:H{row-1})').font=BASE
        ws.cell(row,8).number_format='0.00'; ws.cell(row,8).alignment=Alignment(horizontal='center')
        ws.cell(row,9,f'=SUM(I{first}:I{row-1})').font=BASE
        ws.cell(row,9).number_format='0.00'; ws.cell(row,9).alignment=Alignment(horizontal='center')
        ws.cell(row,15,f'=SUM(O{first}:O{row-1})').font=RED
        ws.cell(row,15).number_format='0'; ws.cell(row,15).alignment=Alignment(horizontal='center')
        ws.cell(row,14,'Meal penalty hours owed, this period ->').font=BOLD
        ws.cell(row,16,f'=IF(J{row}>40,"OVER 40","")').font=RED
        ws.cell(row,16).alignment=Alignment(horizontal='center')
        held_n=sum(1 for x in recs if x['review'])
        if held_n: ws.cell(row,NCOL,f'{held_n} day(s) held - total is incomplete').font=RED
        for col in range(1,NCOL+1):
            ws.cell(row,col).fill=TOT_F; ws.cell(row,col).border=BOX
        row+=2
    ws.sheet_view.showGridLines=False

# ── NEEDS REVIEW ──────────────────────────────────────────────
ws=wb.create_sheet('Needs Review')
ws['A1']='Shifts held back — correct these before paying'; ws['A1'].font=TITLE
ws['A2']=('Every row here is excluded from the pay period totals. A span over 16 hours, or a shift with no '
          'clock-out at all, is treated as a missed punch rather than time worked.'); ws['A2'].font=MUTED
RC=['Employee','Date','Clock In','Clock Out','Recorded Span (hrs)','Problem','Corrected Clock In','Corrected Clock Out']
head(ws,RC,4,[22,11,20,20,18,40,20,20])
ws.freeze_panes='A5'
row=5
bad=[]
for x in D:
    for s in x['segs']:
        if s['bad']: bad.append((x['name'],x['date'],iso(s['i']),iso(s['o']),s['hrs'],s['why']))
for name,d,i,o,h,why in sorted(bad,key=lambda b:(b[0],b[2])):
    ws.cell(row,1,name).font=BASE
    c=ws.cell(row,2,d); c.font=BASE; c.number_format='mm/dd/yyyy'
    c=ws.cell(row,3,i); c.font=INPUT; c.number_format='mm/dd/yyyy h:mm AM/PM'
    if o:
        c=ws.cell(row,4,o); c.font=INPUT; c.number_format='mm/dd/yyyy h:mm AM/PM'
        c=ws.cell(row,5,round(h,2)); c.number_format='0.00'
    else:
        ws.cell(row,4,'— none —').font=RED
    ws.cell(row,5).alignment=Alignment(horizontal='center')
    ws.cell(row,6,why).font=BASE
    for col in range(1,9):
        ws.cell(row,col).border=BOX
        if col<7: ws.cell(row,col).fill=HOLD_F
    row+=1
ws.cell(row+1,1,f'{len(bad)} shifts held').font=BOLD
ws.sheet_view.showGridLines=False

# ── ALL PUNCHES ───────────────────────────────────────────────
ws=wb.create_sheet('All Punches')
ws['A1']='Every punch recorded, unmodified'; ws['A1'].font=TITLE
AC=['Employee','Role','Location','Clock In','Clock Out','Span (hrs)','Lunch Clock-Out','Status']
head(ws,AC,3,[22,14,12,22,22,12,14,14])
ws.freeze_panes='A4'
row=4
for pid,eid,name,role,loc,i,o,lo in sorted(P,key=lambda p:(p[2],p[5])):
    i,o=iso(i),iso(o)
    ws.cell(row,1,name).font=BASE; ws.cell(row,2,role).font=BASE; ws.cell(row,3,loc).font=BASE
    c=ws.cell(row,4,i); c.font=INPUT; c.number_format='mm/dd/yyyy h:mm AM/PM'
    if o:
        c=ws.cell(row,5,o); c.font=INPUT; c.number_format='mm/dd/yyyy h:mm AM/PM'
        ws.cell(row,6,f'=(E{row}-D{row})*24').number_format='0.00'
        ok=(o-i).total_seconds()/3600<=16
    else: ok=False
    ws.cell(row,6).alignment=Alignment(horizontal='center')
    ws.cell(row,7,'Yes' if lo else '').alignment=Alignment(horizontal='center')
    ws.cell(row,8,'OK' if ok else 'HELD').font=BASE if ok else RED
    ws.cell(row,8).alignment=Alignment(horizontal='center')
    if not ok:
        for col in range(1,9): ws.cell(row,col).fill=HOLD_F
    for col in range(1,9): ws.cell(row,col).border=BOX
    row+=1
ws.sheet_view.showGridLines=False

for s in wb: s.sheet_properties.tabColor='1F3864' if s.title=='Read Me' else None
wb.save('Haunted-Trail-Hours-Aug-2026.xlsx')
print('written:', wb.sheetnames)
